import csv
import os
import shutil
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

current_directory = os.path.dirname(os.path.abspath(__file__))
pdf = current_directory + '/tmp/selene-readthedocs-io-en-latest.pdf'
xlsx = current_directory + '/tmp/file_example_XLSX_50.xlsx'
csv_ = current_directory + '/tmp/username.csv'


def create_zip(path):
    zip_ = ZipFile(path, 'w')
    zip_.write(pdf, os.path.basename(pdf))
    zip_.write(xlsx, os.path.basename(xlsx))
    zip_.write(csv_, os.path.basename(csv_))
    zip_.close()
    return shutil.rmtree(current_directory + '/tmp')


def extract_zip_to(to):
    zip_ = ZipFile(current_directory + '/resources/archive.zip')
    zip_.extractall(to)
    zip_.close()


def read_csv(path):
    with open(path) as csvfile:
        csvfile = csv.reader(csvfile)
        return str(csvfile.dialect.__sizeof__())


def read_xlsx(path):
    workbook = load_workbook(path)
    sheet = workbook.active
    return sheet.cell(row=2, column=2).value


def read_pdf(path):
    reader = PdfReader(path)
    page = reader.pages[0]
    return page.extract_text()
